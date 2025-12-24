#!/usr/bin/env python3
"""
npb_boxscores_to_retrosheet_csv.py

Convert standardized 2689web-derived NPB boxscore XLSX files (single sheet, sectioned)
into a Retrosheet-inspired set of CSVs:

- games.csv              : 1 row per game (date/park/attendance/teams/runs/innings, etc.)
- batting.csv            : 1 row per player-game batting line
- pitching.csv           : 1 row per pitcher-game line
- xbh.csv                : extra-base hits (2B/3B) by player-game
- hr_events.csv          : HR events with batter/pitcher attribution when present

Assumptions (based on your standardized output):
- One worksheet named "Sheet" (or first sheet).
- Section headers appear in column A, e.g.:
  "Game Info", "Linescore", "Visitor Batting", "Home Batting",
  "Visitor Pitching", "Home Pitching", "Home Runs", "Extra-Base Hits"
- Batting tables use TWO header rows:
  Row1: Pos, Sub, Name, AB, H
  Row2: R, K, BB, SB, E, AVG, HR
- "Home Runs" section provides team + details string like:
    伊藤裕1号(加藤貴)、フランコ1号(加藤貴)
- "Extra-Base Hits" section provides Visitor/Home, Type (2B/3B), Players list.
"""

from __future__ import annotations

import argparse
import csv
import os
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl


# -------------------------
# Helpers
# -------------------------

def norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()

def split_japanese_list(s: str) -> List[str]:
    """Split on common Japanese separators for name lists."""
    if not s:
        return []
    parts = re.split(r"[、,，;；/／]", s)
    return [p.strip() for p in parts if p and p.strip()]

def parse_players_with_optional_count(s: str) -> List[Tuple[str, int]]:
    """
    Parse a players list cell like:
      '阿部、フランコ' -> [('阿部',1), ('フランコ',1)]
      '山田2、鈴木'    -> [('山田',2), ('鈴木',1)]   (rare but supported)
    """
    s = norm(s)
    if not s or s == "なし":
        return []
    out: List[Tuple[str, int]] = []
    for token in split_japanese_list(s):
        m = re.match(r"^(?P<name>.+?)(?P<count>\d+)$", token)
        if m:
            out.append((m.group("name"), int(m.group("count"))))
        else:
            out.append((token, 1))
    return out

def parse_game_info_jp(game_info: str) -> Dict[str, Any]:
    """
    Best-effort parse of a line like:
      '3月30日　1回戦　エスコンフィールド北海道　31,092人'
    Returns dict with:
      month, day, game_no (optional), park (optional), attendance (optional)
    """
    game_info = norm(game_info)
    res: Dict[str, Any] = {"raw": game_info}

    # month/day
    md = re.search(r"(\d{1,2})月(\d{1,2})日", game_info)
    if md:
        res["month"] = int(md.group(1))
        res["day"] = int(md.group(2))

    # game number e.g. 1回戦
    gno = re.search(r"(\d+)回戦", game_info)
    if gno:
        res["game_no"] = int(gno.group(1))

    # attendance like 31,092人
    att = re.search(r"([\d,]+)人", game_info)
    if att:
        res["attendance"] = int(att.group(1).replace(",", ""))

    # park: remove parsed bits and split by whitespace
    # Heuristic: park is the token before attendance, after game_no if present.
    tokens = re.split(r"\s+|　+", game_info)  # normal + fullwidth spaces
    tokens = [t for t in tokens if t]
    # Example tokens: ['3月30日','1回戦','エスコンフィールド北海道','31,092人']
    if len(tokens) >= 3:
        # park is often tokens[-2]
        park_candidate = tokens[-2]
        if "人" not in park_candidate and "回戦" not in park_candidate and "月" not in park_candidate:
            res["park"] = park_candidate

    return res

def make_game_id(source_filename: str) -> str:
    """
    Derive a stable game_id from filename. Example:
      eagles_2023_EF1.xlsx -> eagles_2023_EF1
    """
    base = Path(source_filename).name
    return base.replace(".xlsx", "")

def safe_int(s: Any) -> Optional[int]:
    s = norm(s)
    if s == "":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None

def safe_float(s: Any) -> Optional[float]:
    s = norm(s)
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


# HR token regex: batter + seasonHRNo + "号" + "(pitcher)"
HR_TOKEN_RE = re.compile(r"^(?P<batter>.+?)(?P<season_no>\d+)号\((?P<pitcher>.+?)\)$")


def parse_hr_details(details: str) -> List[Dict[str, Any]]:
    """
    Parse HR details string into events.
    Example: '伊藤裕1号(加藤貴)、フランコ1号(加藤貴)'
    Returns list of dicts: batter, pitcher, season_no
    """
    details = norm(details)
    if not details or details == "なし":
        return []
    events: List[Dict[str, Any]] = []
    for token in split_japanese_list(details):
        token = token.replace(" ", "").replace("\u3000", "")
        m = HR_TOKEN_RE.match(token)
        if not m:
            # keep raw if unparseable
            events.append({"raw": token})
            continue
        events.append({
            "batter": m.group("batter"),
            "pitcher": m.group("pitcher"),
            "season_no": int(m.group("season_no")),
        })
    return events


# -------------------------
# Section parsing
# -------------------------

@dataclass
class SectionIndex:
    game_info_row: Optional[int] = None
    linescore_row: Optional[int] = None
    vis_bat_row: Optional[int] = None
    home_bat_row: Optional[int] = None
    vis_pitch_row: Optional[int] = None
    home_pitch_row: Optional[int] = None
    hr_row: Optional[int] = None
    xbh_row: Optional[int] = None


def read_sheet_matrix(ws) -> List[List[Any]]:
    """Read the used range into a matrix (small sheets)."""
    max_row = ws.max_row
    max_col = ws.max_column
    mat: List[List[Any]] = []
    for r in range(1, max_row + 1):
        mat.append([ws.cell(r, c).value for c in range(1, max_col + 1)])
    return mat


def find_sections(mat: List[List[Any]]) -> SectionIndex:
    idx = SectionIndex()
    for i, row in enumerate(mat, start=1):
        a = norm(row[0]) if row else ""
        if a == "Game Info":
            idx.game_info_row = i
        elif a == "Linescore":
            idx.linescore_row = i
        elif a == "Visitor Batting":
            idx.vis_bat_row = i
        elif a == "Home Batting":
            idx.home_bat_row = i
        elif a == "Visitor Pitching":
            idx.vis_pitch_row = i
        elif a == "Home Pitching":
            idx.home_pitch_row = i
        elif a == "Home Runs":
            idx.hr_row = i
        elif a == "Extra-Base Hits":
            idx.xbh_row = i
    return idx


def parse_linescore(mat: List[List[Any]], start_row: int) -> Dict[str, Any]:
    """
    Parse linescore table that begins right after a 'Linescore' row.
    Expected structure:
      Row+1: TEAM | 1 | 2 | ... | 9 (or more)
      Row+2: AWAY_TEAM | runs by inning...
      Row+3: HOME_TEAM | runs by inning...
    """
    header = mat[start_row] if start_row < len(mat) else None
    # We actually want row after 'Linescore' marker (start_row+1 in 1-based terms)
    r_team_header = start_row + 1
    r_away = start_row + 2
    r_home = start_row + 3

    def rowvals(r: int) -> List[str]:
        if r < 1 or r > len(mat):
            return []
        return [norm(x) for x in mat[r - 1]]

    h = rowvals(r_team_header)
    away = rowvals(r_away)
    home = rowvals(r_home)

    # inning columns are whatever non-empty headers after TEAM
    inning_headers = [x for x in h[1:] if x != ""]
    away_innings = [safe_int(x) for x in away[1:1+len(inning_headers)]]
    home_innings = [safe_int(x) for x in home[1:1+len(inning_headers)]]

    def sum_innings(vals: List[Optional[int]]) -> int:
        return sum(v for v in vals if isinstance(v, int))

    return {
        "away_team": away[0],
        "home_team": home[0],
        "inning_headers": inning_headers,
        "away_by_inning": away_innings,
        "home_by_inning": home_innings,
        "away_runs": sum_innings(away_innings),
        "home_runs": sum_innings(home_innings),
        "innings": len(inning_headers),
    }


def parse_batting_table(mat: List[List[Any]], start_row: int) -> List[Dict[str, Any]]:
    """
    Parse a batting table starting after 'Visitor Batting' or 'Home Batting'.

    Expected:
      start_row+1: Pos | Sub | Name | AB | H | ...
      start_row+? : second header row containing R,K,BB,SB,E,AVG,HR
      Then player rows until blank row.
    """
    # locate first non-empty header row after marker
    r1 = start_row + 1
    # find first row where col0 == 'Pos'
    while r1 <= len(mat) and norm(mat[r1-1][0]) != "Pos":
        r1 += 1
    if r1 > len(mat):
        return []

    # header row 1 (Pos/Sub/Name/AB/H)
    header1 = [norm(x) for x in mat[r1-1]]

    # header row 2 is usually a few rows later where col0 == 'R'
    r2 = r1 + 1
    while r2 <= len(mat) and norm(mat[r2-1][0]) != "R":
        r2 += 1
    if r2 > len(mat):
        return []

    header2 = [norm(x) for x in mat[r2-1]]

    # Combine headers by position:
    # first columns from header1 until blanks, then header2 continues
    # In your sample, header1: Pos, Sub, Name, AB, H
    # header2: R, K, BB, SB, E, AVG, HR
    columns = []
    # take non-empty from header1
    for h in header1:
        if h:
            columns.append(h)
    # take non-empty from header2
    for h in header2:
        if h:
            columns.append(h)

    # data starts at r2+1
    out: List[Dict[str, Any]] = []
    r = r2 + 1
    while r <= len(mat):
        row = [norm(x) for x in mat[r-1]]
        if all(x == "" for x in row):
            break
        # stop if we hit another section marker in col0
        if row[0] in {"Home Batting", "Visitor Pitching", "Home Pitching", "Home Runs", "Extra-Base Hits", "Linescore"}:
            break

        # Build record by reading across: first chunk is header1 length, second chunk is header2 length
        # We use the original table layout: row has Pos/Sub/Name/AB/H then later R/K/BB/... etc.
        # So map by known indexes:
        rec = {
            "Pos": row[0],
            "Sub": row[1],
            "Name": row[2],
            "AB": safe_int(row[3]),
            "H": safe_int(row[4]),
            "R": safe_int(row[5]) if len(row) > 5 else None,
            "K": safe_int(row[6]) if len(row) > 6 else None,
            "BB": safe_int(row[7]) if len(row) > 7 else None,
            "SB": safe_int(row[8]) if len(row) > 8 else None,
            "E": safe_int(row[9]) if len(row) > 9 else None,
            "AVG": row[10] if len(row) > 10 else "",
            "HR": safe_int(row[11]) if len(row) > 11 else None,
        }

        # Skip obvious non-player subtotal rows (often Name blank)
        if rec["Name"]:
            out.append(rec)

        r += 1

    return out


def parse_pitching_table(mat: List[List[Any]], start_row: int) -> List[Dict[str, Any]]:
    """
    Parse pitching table starting after 'Visitor Pitching'/'Home Pitching'.

    We assume your standardized sheet already has pitching rows aligned like:
      Name | IP | BF | H | SO | BB | ER | R | WLS(season) | ERA(season)
    We'll read rows until blank.
    """
    r = start_row + 1
    # Skip blank spacer rows
    while r <= len(mat) and all(norm(x) == "" for x in mat[r-1]):
        r += 1

    out: List[Dict[str, Any]] = []
    while r <= len(mat):
        row = [norm(x) for x in mat[r-1]]
        if all(x == "" for x in row):
            break
        if row[0] in {"Home Pitching", "Home Runs", "Extra-Base Hits"}:
            break

        # Expected columns present in first 10
        rec = {
            "Name": row[0],
            "IP": safe_float(row[1]),
            "BF": safe_int(row[2]),
            "H": safe_int(row[3]),
            "SO": safe_int(row[4]),
            "BB": safe_int(row[5]),
            "ER": safe_int(row[6]),
            "R": safe_int(row[7]),
            "WLS_season": row[8] if len(row) > 8 else "",
            "ERA_season": safe_float(row[9]) if len(row) > 9 else None,
        }
        if rec["Name"]:
            out.append(rec)
        r += 1

    return out


def parse_hr_section(mat: List[List[Any]], start_row: int) -> List[Dict[str, Any]]:
    """
    Parse Home Runs section:
      Row: 'Home Runs'
      Row+1: 'Team' | 'Details'
      Row+2+: team | details until blank
    """
    r = start_row + 1
    # find header row that starts with 'Team'
    while r <= len(mat) and norm(mat[r-1][0]) != "Team":
        r += 1
    if r > len(mat):
        return []

    r += 1
    out: List[Dict[str, Any]] = []
    while r <= len(mat):
        row = [norm(x) for x in mat[r-1]]
        if all(x == "" for x in row):
            break
        team = row[0]
        details = row[1] if len(row) > 1 else ""
        out.append({"team": team, "details": details, "events": parse_hr_details(details)})
        r += 1
    return out


def parse_xbh_section(mat: List[List[Any]], start_row: int) -> List[Dict[str, Any]]:
    """
    Parse Extra-Base Hits section:
      Row: 'Extra-Base Hits'
      Row+1: 'Team' | 'Type' | 'Players'
      Then rows with Team in {Visitor, Home} and Type in {2B,3B}
    """
    r = start_row + 1
    while r <= len(mat) and norm(mat[r-1][0]) != "Team":
        r += 1
    if r > len(mat):
        return []

    r += 1
    out: List[Dict[str, Any]] = []
    while r <= len(mat):
        row = [norm(x) for x in mat[r-1]]
        if all(x == "" for x in row):
            break
        team_side = row[0]
        xbh_type = row[1] if len(row) > 1 else ""
        players = row[2] if len(row) > 2 else ""
        out.append({"team_side": team_side, "type": xbh_type, "players": players, "parsed": parse_players_with_optional_count(players)})
        r += 1
    return out


# -------------------------
# Main conversion
# -------------------------

def convert_file(xlsx_path: Path) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    mat = read_sheet_matrix(ws)
    sec = find_sections(mat)

    game_id = make_game_id(xlsx_path.name)

    # game info
    gi_value = ""
    if sec.game_info_row:
        # the value should be in column B of the 'Game Info' row
        row = mat[sec.game_info_row - 1]
        gi_value = norm(row[1]) if len(row) > 1 else ""
    gi = parse_game_info_jp(gi_value)

    # linescore
    linescore = {}
    if sec.linescore_row:
        linescore = parse_linescore(mat, sec.linescore_row)

    # batting
    vis_bat = parse_batting_table(mat, sec.vis_bat_row) if sec.vis_bat_row else []
    home_bat = parse_batting_table(mat, sec.home_bat_row) if sec.home_bat_row else []

    # pitching
    vis_pitch = parse_pitching_table(mat, sec.vis_pitch_row) if sec.vis_pitch_row else []
    home_pitch = parse_pitching_table(mat, sec.home_pitch_row) if sec.home_pitch_row else []

    # HR + XBH
    hr = parse_hr_section(mat, sec.hr_row) if sec.hr_row else []
    xbh = parse_xbh_section(mat, sec.xbh_row) if sec.xbh_row else []

    return {
        "game_id": game_id,
        "source_file": str(xlsx_path),
        "game_info": gi,
        "linescore": linescore,
        "visitor_batting": vis_bat,
        "home_batting": home_bat,
        "visitor_pitching": vis_pitch,
        "home_pitching": home_pitch,
        "home_runs": hr,
        "extra_base_hits": xbh,
    }


def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, "") for k in fieldnames})


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input-dir", required=True, help="Folder containing standardized boxscore XLSX files")
    ap.add_argument("--output-dir", required=True, help="Folder to write CSV outputs")
    ap.add_argument("--glob", default="*.xlsx", help="Glob for input files (default: *.xlsx)")
    args = ap.parse_args()

    in_dir = Path(args.input_dir)
    out_dir = Path(args.output_dir)
    files = sorted(in_dir.glob(args.glob))

    games_rows: List[Dict[str, Any]] = []
    batting_rows: List[Dict[str, Any]] = []
    pitching_rows: List[Dict[str, Any]] = []
    xbh_rows: List[Dict[str, Any]] = []
    hr_event_rows: List[Dict[str, Any]] = []

    for fp in files:
        rec = convert_file(fp)
        gid = rec["game_id"]
        gi = rec["game_info"]
        ls = rec.get("linescore", {}) or {}

        games_rows.append({
            "game_id": gid,
            "source_file": rec["source_file"],
            "month": gi.get("month", ""),
            "day": gi.get("day", ""),
            "game_no": gi.get("game_no", ""),
            "park": gi.get("park", ""),
            "attendance": gi.get("attendance", ""),
            "away_team": ls.get("away_team", ""),
            "home_team": ls.get("home_team", ""),
            "away_runs": ls.get("away_runs", ""),
            "home_runs": ls.get("home_runs", ""),
            "innings": ls.get("innings", ""),
        })

        for side, rows in (("Visitor", rec["visitor_batting"]), ("Home", rec["home_batting"])):
            for r in rows:
                batting_rows.append({
                    "game_id": gid,
                    "team_side": side,
                    "pos": r.get("Pos", ""),
                    "sub": r.get("Sub", ""),
                    "player_name": r.get("Name", ""),
                    "ab": r.get("AB", ""),
                    "h": r.get("H", ""),
                    "r": r.get("R", ""),
                    "so": r.get("K", ""),
                    "bb": r.get("BB", ""),
                    "sb": r.get("SB", ""),
                    "e": r.get("E", ""),
                    "avg_season": r.get("AVG", ""),
                    "hr_game": r.get("HR", ""),
                })

        for side, rows in (("Visitor", rec["visitor_pitching"]), ("Home", rec["home_pitching"])):
            for r in rows:
                pitching_rows.append({
                    "game_id": gid,
                    "team_side": side,
                    "pitcher_name": r.get("Name", ""),
                    "ip": r.get("IP", ""),
                    "bf": r.get("BF", ""),
                    "h": r.get("H", ""),
                    "so": r.get("SO", ""),
                    "bb": r.get("BB", ""),
                    "er": r.get("ER", ""),
                    "r": r.get("R", ""),
                    "wls_season": r.get("WLS_season", ""),
                    "era_season": r.get("ERA_season", ""),
                })

        # XBH rows into per-player rows
        for x in rec["extra_base_hits"]:
            for name, cnt in x.get("parsed", []):
                xbh_rows.append({
                    "game_id": gid,
                    "team_side": x.get("team_side", ""),
                    "xbh_type": x.get("type", ""),
                    "player_name": name,
                    "count": cnt,
                })

        # HR events
        for t in rec["home_runs"]:
            for ev in t.get("events", []):
                if "batter" in ev:
                    hr_event_rows.append({
                        "game_id": gid,
                        "team_label": t.get("team", ""),
                        "batter": ev.get("batter", ""),
                        "pitcher": ev.get("pitcher", ""),
                        "season_no": ev.get("season_no", ""),
                        "raw": "",
                    })
                else:
                    hr_event_rows.append({
                        "game_id": gid,
                        "team_label": t.get("team", ""),
                        "batter": "",
                        "pitcher": "",
                        "season_no": "",
                        "raw": ev.get("raw", ""),
                    })

    write_csv(out_dir / "games.csv", games_rows,
              ["game_id","source_file","month","day","game_no","park","attendance",
               "away_team","home_team","away_runs","home_runs","innings"])

    write_csv(out_dir / "batting.csv", batting_rows,
              ["game_id","team_side","pos","sub","player_name","ab","h","r","so","bb","sb","e","avg_season","hr_game"])

    write_csv(out_dir / "pitching.csv", pitching_rows,
              ["game_id","team_side","pitcher_name","ip","bf","h","so","bb","er","r","wls_season","era_season"])

    write_csv(out_dir / "xbh.csv", xbh_rows,
              ["game_id","team_side","xbh_type","player_name","count"])

    write_csv(out_dir / "hr_events.csv", hr_event_rows,
              ["game_id","team_label","batter","pitcher","season_no","raw"])

    print(f"Converted {len(files)} files -> {out_dir}")


if __name__ == "__main__":
    main()
