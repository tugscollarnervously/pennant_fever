#!/usr/bin/env python3
"""
npb_boxscore_converter.py

Convert standardized 2689web NPB boxscore XLSX files into normalized CSV files
suitable for SQL database storage and querying.

Output CSVs:
- games.csv          : One row per game (metadata + final score)
- linescore.csv      : One row per team per inning
- batting.csv        : One row per batter appearance per game
- pitching.csv       : One row per pitcher appearance per game
- home_runs.csv      : One row per HR event
- extra_base_hits.csv: One row per 2B/3B event
"""

import argparse
import csv
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


# =============================================================================
# STANDARDIZATION MAPPINGS
# =============================================================================

# Japanese position codes to standard baseball abbreviations
POSITION_MAP = {
    '投': 'P',
    '捕': 'C',
    '一': '1B',
    '二': '2B',
    '三': '3B',
    '遊': 'SS',
    '左': 'LF',
    '中': 'CF',
    '右': 'RF',
    '指': 'DH',
    # Compound positions (defensive subs)
    '左中': 'LF-CF',
    '右中': 'RF-CF',
    '走左': 'PR-LF',
    '走右': 'PR-RF',
    '走中': 'PR-CF',
    '走一': 'PR-1B',
    '走二': 'PR-2B',
    '走三': 'PR-3B',
    '走遊': 'PR-SS',
}

# Sub column codes
SUB_TYPE_MAP = {
    '打': 'PH',      # Pinch hitter
    '走': 'PR',      # Pinch runner
    '捕': 'C',       # Defensive sub at catcher
    '一': '1B',
    '二': '2B',
    '三': '3B',
    '遊': 'SS',
    '左': 'LF',
    '中': 'CF',
    '右': 'RF',
}

# Decision codes
DECISION_MAP = {
    '勝': 'W',
    '敗': 'L',
    'Ｓ': 'SV',
    'S': 'SV',
    'Ｈ': 'H',
    'H': 'H',
}

# Team URL codes to franchise codes
TEAM_MAP = {
    'BRAVES': 'BRV',
    'BLUEWAVE': 'BW',
    'BUFFALOES': 'BUF',
    'CARP': 'CAR',
    'CLIPPERS': 'CLI',
    'LIONS': 'LIO',
    'DRAGONS': 'DRA',
    'EAGLES': 'EAG',
    'GIANTS': 'GIA',
    'ORIONS': 'ORI',
    'PIRATES': 'PIR',
    'SWALLOWS': 'SWA',
    'TIGERS': 'TIG',
    'FIGHTERS': 'FIG',
    'HAWKS': 'HAW',
    'MARINES': 'MAR',
    'BAYSTARS': 'BAY',
    'WHALES': 'WHA',
    # Handle hyphenated versions from linescore (e.g., NIPPONHAM-5)
    'NIPPONHAM': 'FIG',
    'SOFTBANK': 'HAW',
    'LOTTE': 'MAR',
    'YOKOHAMA': 'BAY',
    'ORIX': 'BUF',
    'RAKUTEN': 'EAG',
    'SEIBU': 'LIO',
    'CHUNICHI': 'DRA',
    'YOMIURI': 'GIA',
    'HANSHIN': 'TIG',
    'HIROSHIMA': 'CAR',
    'YAKULT': 'SWA',
    'KINTETSU': 'BUF',
    'HANKYU': 'BRV',
    'NANKAI': 'HAW',
    'NISHITETSU': 'LIO',
    'TAIYO': 'BAY',
    'TAIHEIYO': 'LIO',
    'CROWN': 'LIO',
    'KYOJIN': 'GIA',
    'KOKUTETSU': 'SWA',
}

# Japanese team names to romanized equivalents (for HR/XBH matching)
TEAM_JP_MAP = {
    '楽天': 'RAKUTEN',
    '日本ハム': 'NIPPONHAM',
    'ソフトバンク': 'SOFTBANK',
    'ロッテ': 'LOTTE',
    'オリックス': 'ORIX',
    '西武': 'SEIBU',
    '巨人': 'KYOJIN',
    '阪神': 'HANSHIN',
    '中日': 'CHUNICHI',
    'ヤクルト': 'YAKULT',
    '広島': 'HIROSHIMA',
    'DeNA': 'YOKOHAMA',
    '横浜': 'YOKOHAMA',
    '南海': 'NANKAI',
    '近鉄': 'KINTETSU',
    '阪急': 'HANKYU',
    '大洋': 'TAIYO',
    '国鉄': 'KOKUTETSU',
    '西鉄': 'NISHITETSU',
}


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def norm(value: Any) -> str:
    """Normalize a cell value to string, stripping whitespace."""
    if value is None:
        return ''
    return str(value).strip()


def safe_int(value: Any) -> Optional[int]:
    """Safely convert to int, return None on failure."""
    s = norm(value)
    if not s:
        return None
    # Handle 'x' suffix for walkoffs (e.g., '1x')
    s = s.rstrip('x').rstrip('X')
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def safe_float(value: Any) -> Optional[float]:
    """Safely convert to float, return None on failure."""
    s = norm(value)
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def extract_year_from_filename(filename: str) -> Optional[int]:
    """Extract year from filename like 'eagles_2023_EF1.xlsx'."""
    match = re.search(r'_(\d{4})_', filename)
    if match:
        return int(match.group(1))
    return None


def extract_team_from_filename(filename: str) -> str:
    """Extract team name from filename like 'eagles_2023_EF1.xlsx'."""
    match = re.match(r'^([a-z]+)_', filename.lower())
    if match:
        return match.group(1).upper()
    return ''


def normalize_team_name(raw_name: str) -> str:
    """Normalize team name from linescore to standard code."""
    # Remove numeric suffixes (e.g., NIPPONHAM-5 -> NIPPONHAM)
    name = re.sub(r'-\d+$', '', raw_name.upper())
    return TEAM_MAP.get(name, name[:3] if name else 'UNK')


def normalize_position(pos: str) -> str:
    """Convert Japanese position to standard abbreviation."""
    pos = norm(pos)
    if not pos:
        return ''
    # Check direct mapping
    if pos in POSITION_MAP:
        return POSITION_MAP[pos]
    # Check if first character matches
    if pos and pos[0] in POSITION_MAP:
        return POSITION_MAP[pos[0]]
    return pos


def normalize_sub_type(sub: str, pos: str = '') -> str:
    """Determine sub type from Sub column."""
    sub = norm(sub)
    if not sub:
        return ''
    # Check direct mapping
    if sub in SUB_TYPE_MAP:
        base = SUB_TYPE_MAP[sub]
        # If it's a position code, this is a defensive sub
        if base not in ('PH', 'PR'):
            return f'DEF-{base}'
        return base
    # Check first character
    if sub and sub[0] in SUB_TYPE_MAP:
        base = SUB_TYPE_MAP[sub[0]]
        if base not in ('PH', 'PR'):
            return f'DEF-{base}'
        return base
    return sub


def normalize_decision(dec: str) -> str:
    """Convert Japanese decision code to W/L/SV/H."""
    dec = norm(dec)
    return DECISION_MAP.get(dec, dec)


def parse_game_info(game_info: str) -> Dict[str, Any]:
    """
    Parse game info string like '4月5日　1回戦　後楽園球場　49,000人'
    Returns dict with month, day, game_number, stadium, attendance.
    """
    info = norm(game_info)
    result: Dict[str, Any] = {'raw': info}

    # Month/day: 4月5日
    md_match = re.search(r'(\d{1,2})月(\d{1,2})日', info)
    if md_match:
        result['month'] = int(md_match.group(1))
        result['day'] = int(md_match.group(2))

    # Game number: 1回戦
    gn_match = re.search(r'(\d+)回戦', info)
    if gn_match:
        result['game_number'] = int(gn_match.group(1))

    # Attendance: 49,000人
    att_match = re.search(r'([\d,]+)人', info)
    if att_match:
        result['attendance'] = int(att_match.group(1).replace(',', ''))

    # Stadium: heuristic - token before attendance, after game number
    tokens = re.split(r'[\s　]+', info)
    tokens = [t for t in tokens if t]
    # Stadium is typically 3rd or 2nd-to-last token
    # Keywords: 球場 (stadium), ドーム (dome), スタジアム (stadium), フィールド (field), パーク (park)
    for token in tokens:
        if '球場' in token or 'ドーム' in token or 'スタジアム' in token or 'フィールド' in token or 'パーク' in token:
            result['stadium'] = token
            break

    return result


def parse_hr_token(token: str) -> Dict[str, Any]:
    """
    Parse HR token like '伊藤裕1号(加藤貴)' into batter, pitcher, season_no.
    """
    token = norm(token).replace(' ', '').replace('\u3000', '')
    # Pattern: batter + number + 号 + (pitcher)
    match = re.match(r'^(.+?)(\d+)号\((.+?)\)$', token)
    if match:
        return {
            'batter': match.group(1),
            'pitcher': match.group(3),
            'season_hr_number': int(match.group(2)),
        }
    return {'raw': token}


def split_japanese_list(s: str) -> List[str]:
    """Split on common Japanese list separators."""
    if not s:
        return []
    parts = re.split(r'[、,，;；/／]', s)
    return [p.strip() for p in parts if p.strip()]


def parse_xbh_player(player_str: str) -> Dict[str, Any]:
    """
    Parse XBH player string like '山崎2' into player name and count.
    '山崎2' means Yamazaki hit 2 doubles/triples.
    Returns dict with player_name and count.
    """
    player_str = norm(player_str)
    if not player_str:
        return {'player_name': '', 'count': 1}

    # Pattern: name + optional trailing number
    match = re.match(r'^(.+?)(\d+)$', player_str)
    if match:
        return {
            'player_name': match.group(1),
            'count': int(match.group(2)),
        }
    return {'player_name': player_str, 'count': 1}


# =============================================================================
# EXCEL PARSING
# =============================================================================

def read_sheet_as_matrix(ws) -> List[List[Any]]:
    """Read worksheet into a 2D list (1-indexed row numbers preserved via list index+1)."""
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    matrix = []
    for r in range(1, max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        matrix.append(row)
    return matrix


def find_section_rows(matrix: List[List[Any]]) -> Dict[str, int]:
    """
    Find the row indices (0-based) for each section header.
    Returns dict mapping section name to row index.
    """
    sections = {}
    section_names = [
        'Game Info', 'Linescore', 'Visitor Batting', 'Home Batting',
        'Visitor Pitching', 'Home Pitching', 'Home Runs', 'Extra-Base Hits'
    ]

    for i, row in enumerate(matrix):
        cell0 = norm(row[0]) if row else ''
        if cell0 in section_names:
            sections[cell0] = i

    return sections


def parse_linescore_section(matrix: List[List[Any]], start_row: int) -> Dict[str, Any]:
    """
    Parse linescore section starting at given row.
    Returns dict with header, away team data, home team data.
    """
    # Row structure after 'Linescore':
    # +1: Header row (TEAM, 1, 2, 3, ..., R, H, E)
    # +2: Away team row
    # +3: Home team row

    header_row = matrix[start_row + 1] if start_row + 1 < len(matrix) else []
    away_row = matrix[start_row + 2] if start_row + 2 < len(matrix) else []
    home_row = matrix[start_row + 3] if start_row + 3 < len(matrix) else []

    header = [norm(x) for x in header_row]
    away = [norm(x) for x in away_row]
    home = [norm(x) for x in home_row]

    # Find indices for R, H, E columns
    try:
        r_idx = header.index('R')
        h_idx = header.index('H')
        e_idx = header.index('E')
    except ValueError:
        r_idx = h_idx = e_idx = -1

    # Inning columns are between TEAM and R
    inning_cols = []
    for i, h in enumerate(header[1:], start=1):
        if h == 'R':
            break
        if h and h not in ('', '…'):
            inning_cols.append((i, h))

    return {
        'header': header,
        'away_team': away[0] if away else '',
        'home_team': home[0] if home else '',
        'innings': [(col_name, safe_int(away[col_idx]), safe_int(home[col_idx]))
                    for col_idx, col_name in inning_cols],
        'away_runs': safe_int(away[r_idx]) if r_idx > 0 and r_idx < len(away) else None,
        'home_runs': safe_int(home[r_idx]) if r_idx > 0 and r_idx < len(home) else None,
        'away_hits': safe_int(away[h_idx]) if h_idx > 0 and h_idx < len(away) else None,
        'home_hits': safe_int(home[h_idx]) if h_idx > 0 and h_idx < len(home) else None,
        'away_errors': safe_int(away[e_idx]) if e_idx > 0 and e_idx < len(away) else None,
        'home_errors': safe_int(home[e_idx]) if e_idx > 0 and e_idx < len(home) else None,
    }


def parse_batting_section(matrix: List[List[Any]], start_row: int, end_markers: List[str]) -> List[Dict[str, Any]]:
    """
    Parse batting section starting at given row.
    Returns list of batter records.
    """
    batters = []
    lineup_slot = 0
    current_starter = True

    # Skip header rows - find first data row
    # Header is: Pos, Sub, Name, AB, H, R, K, BB, SB, E, AVG, HR
    r = start_row + 1

    # Skip until we find the header row with 'Pos'
    while r < len(matrix):
        if norm(matrix[r][0]) == 'Pos':
            r += 1  # Move past header
            break
        r += 1

    while r < len(matrix):
        row = [norm(x) for x in matrix[r]]

        # Stop conditions
        if all(x == '' for x in row):
            r += 1
            continue
        if row[0] in end_markers:
            break

        # Skip team totals row (usually has blank Pos and Name with just numbers)
        # Also skip if it looks like a header row
        pos = row[0] if len(row) > 0 else ''
        sub = row[1] if len(row) > 1 else ''
        name = row[2] if len(row) > 2 else ''

        # Skip non-player rows
        if not name or name in ('Pos', 'Sub', 'Name', 'TEAM'):
            r += 1
            continue

        # Determine if starter or sub
        is_starter = (sub == '')
        if is_starter:
            lineup_slot += 1
            current_starter = True

        # Parse stats
        batter = {
            'lineup_slot': lineup_slot,
            'player_name_jp': name,
            'position': normalize_position(pos),
            'is_starter': 1 if is_starter else 0,
            'sub_type': normalize_sub_type(sub, pos) if not is_starter else '',
            'ab': safe_int(row[3]) if len(row) > 3 else None,
            'h': safe_int(row[4]) if len(row) > 4 else None,
            'r': safe_int(row[5]) if len(row) > 5 else None,
            'k': safe_int(row[6]) if len(row) > 6 else None,
            'bb': safe_int(row[7]) if len(row) > 7 else None,
            'sb': safe_int(row[8]) if len(row) > 8 else None,
            'e': safe_int(row[9]) if len(row) > 9 else None,
            'hr': safe_int(row[11]) if len(row) > 11 else None,  # Skip AVG at index 10
        }

        # Only add if has valid AB (filters out totals rows)
        if batter['ab'] is not None:
            batters.append(batter)

        r += 1

    return batters


def parse_pitching_section(matrix: List[List[Any]], start_row: int, end_markers: List[str]) -> List[Dict[str, Any]]:
    """
    Parse pitching section starting at given row.
    Returns list of pitcher records.
    """
    pitchers = []
    pitch_order = 0

    # Skip to header row with 'Dec'
    r = start_row + 1
    while r < len(matrix):
        if norm(matrix[r][0]) == 'Dec':
            r += 1  # Move past header
            break
        r += 1

    while r < len(matrix):
        row = [norm(x) for x in matrix[r]]

        # Stop conditions
        if all(x == '' for x in row):
            r += 1
            continue
        if row[0] in end_markers:
            break

        dec = row[0] if len(row) > 0 else ''
        name = row[1] if len(row) > 1 else ''

        # Skip non-pitcher rows (headers, totals)
        if not name or name in ('Dec', 'Name', 'TEAM'):
            r += 1
            continue

        pitch_order += 1

        pitcher = {
            'pitch_order': pitch_order,
            'player_name_jp': name,
            'decision': normalize_decision(dec),
            'ip': safe_float(row[2]) if len(row) > 2 else None,
            'bf': safe_int(row[3]) if len(row) > 3 else None,
            'h': safe_int(row[4]) if len(row) > 4 else None,
            'k': safe_int(row[5]) if len(row) > 5 else None,
            'bb': safe_int(row[6]) if len(row) > 6 else None,
            'er': safe_int(row[7]) if len(row) > 7 else None,
            'e': safe_int(row[8]) if len(row) > 8 else None,
            # Skip W-L-S and ERA (seasonal stats)
        }

        # Only add if has valid IP
        if pitcher['ip'] is not None:
            pitchers.append(pitcher)

        r += 1

    return pitchers


def parse_hr_section(matrix: List[List[Any]], start_row: int) -> List[Dict[str, Any]]:
    """
    Parse Home Runs section.
    Returns list of HR events.
    """
    events = []

    # Find header row with 'Team'
    r = start_row + 1
    while r < len(matrix):
        if norm(matrix[r][0]) == 'Team':
            r += 1  # Move past header
            break
        r += 1

    while r < len(matrix):
        row = [norm(x) for x in matrix[r]]

        if all(x == '' for x in row):
            break
        if row[0] in ('Extra-Base Hits', 'No home runs'):
            break

        team = row[0] if len(row) > 0 else ''
        details = row[1] if len(row) > 1 else ''

        if not team or details in ('なし', 'No home runs', ''):
            r += 1
            continue

        # Parse individual HR tokens from details
        for token in split_japanese_list(details):
            hr_data = parse_hr_token(token)
            hr_data['team_raw'] = team
            events.append(hr_data)

        r += 1

    return events


def parse_xbh_section(matrix: List[List[Any]], start_row: int) -> List[Dict[str, Any]]:
    """
    Parse Extra-Base Hits section.
    Returns list of XBH events.
    """
    events = []

    # Find header row with 'Team'
    r = start_row + 1
    while r < len(matrix):
        if norm(matrix[r][0]) == 'Team':
            r += 1  # Move past header
            break
        r += 1

    while r < len(matrix):
        row = [norm(x) for x in matrix[r]]

        if all(x == '' for x in row):
            break
        if row[0] in ('No extra-base hits',):
            break

        team_side = row[0] if len(row) > 0 else ''  # 'Visitor' or 'Home'
        hit_type = row[1] if len(row) > 1 else ''   # '2B' or '3B'
        players = row[2] if len(row) > 2 else ''

        if not team_side or players in ('なし', ''):
            r += 1
            continue

        # Split players and create individual events
        for player in split_japanese_list(players):
            events.append({
                'team_side': team_side,
                'hit_type': hit_type,
                'player_name_jp': player,
            })

        r += 1

    return events


# =============================================================================
# MAIN CONVERSION
# =============================================================================

def convert_boxscore(xlsx_path: Path) -> Dict[str, Any]:
    """
    Convert a single boxscore XLSX file to structured data.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    matrix = read_sheet_as_matrix(ws)
    sections = find_section_rows(matrix)

    # Extract metadata from filename
    filename = xlsx_path.name
    game_id = filename.replace('.xlsx', '')
    year = extract_year_from_filename(filename)
    team_from_file = extract_team_from_filename(filename)

    # Parse game info
    game_info = {}
    if 'Game Info' in sections:
        row = matrix[sections['Game Info']]
        raw_info = norm(row[1]) if len(row) > 1 else ''
        game_info = parse_game_info(raw_info)

    # Parse linescore
    linescore = {}
    if 'Linescore' in sections:
        linescore = parse_linescore_section(matrix, sections['Linescore'])

    # Define end markers for sections
    batting_end = ['Home Batting', 'Visitor Pitching', 'Home Pitching', 'Home Runs', 'Extra-Base Hits']
    pitching_end = ['Home Pitching', 'Home Runs', 'Extra-Base Hits']

    # Parse batting
    visitor_batting = []
    if 'Visitor Batting' in sections:
        visitor_batting = parse_batting_section(matrix, sections['Visitor Batting'], batting_end)

    home_batting = []
    if 'Home Batting' in sections:
        home_batting = parse_batting_section(matrix, sections['Home Batting'], batting_end)

    # Parse pitching
    visitor_pitching = []
    if 'Visitor Pitching' in sections:
        visitor_pitching = parse_pitching_section(matrix, sections['Visitor Pitching'], pitching_end)

    home_pitching = []
    if 'Home Pitching' in sections:
        home_pitching = parse_pitching_section(matrix, sections['Home Pitching'], ['Home Runs', 'Extra-Base Hits'])

    # Parse home runs
    home_runs = []
    if 'Home Runs' in sections:
        home_runs = parse_hr_section(matrix, sections['Home Runs'])

    # Parse extra-base hits
    xbh = []
    if 'Extra-Base Hits' in sections:
        xbh = parse_xbh_section(matrix, sections['Extra-Base Hits'])

    wb.close()

    return {
        'game_id': game_id,
        'year': year,
        'game_info': game_info,
        'linescore': linescore,
        'visitor_batting': visitor_batting,
        'home_batting': home_batting,
        'visitor_pitching': visitor_pitching,
        'home_pitching': home_pitching,
        'home_runs': home_runs,
        'extra_base_hits': xbh,
    }


def write_csvs(games_data: List[Dict[str, Any]], output_dir: Path):
    """
    Write all parsed game data to CSV files.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    # Collect rows for each CSV
    games_rows = []
    linescore_rows = []
    batting_rows = []
    pitching_rows = []
    hr_rows = []
    xbh_rows = []

    for game in games_data:
        gid = game['game_id']
        year = game['year']
        gi = game.get('game_info', {})
        ls = game.get('linescore', {})

        # Build date string
        month = gi.get('month', '')
        day = gi.get('day', '')
        date_str = f"{year}-{month:02d}-{day:02d}" if year and month and day else ''

        # Games row
        games_rows.append({
            'game_id': gid,
            'year': year or '',
            'month': month,
            'day': day,
            'date': date_str,
            'game_number': gi.get('game_number', ''),
            'stadium': gi.get('stadium', ''),
            'attendance': gi.get('attendance', ''),
            'away_team': normalize_team_name(ls.get('away_team', '')),
            'home_team': normalize_team_name(ls.get('home_team', '')),
            'away_team_raw': ls.get('away_team', ''),
            'home_team_raw': ls.get('home_team', ''),
            'away_runs': ls.get('away_runs', ''),
            'home_runs': ls.get('home_runs', ''),
            'away_hits': ls.get('away_hits', ''),
            'home_hits': ls.get('home_hits', ''),
            'away_errors': ls.get('away_errors', ''),
            'home_errors': ls.get('home_errors', ''),
            'innings_played': len(ls.get('innings', [])),
        })

        # Linescore rows (one per team per inning)
        for inning_name, away_runs, home_runs in ls.get('innings', []):
            if away_runs is not None:
                linescore_rows.append({
                    'game_id': gid,
                    'team_side': 'away',
                    'inning': inning_name,
                    'runs': away_runs,
                })
            if home_runs is not None:
                linescore_rows.append({
                    'game_id': gid,
                    'team_side': 'home',
                    'inning': inning_name,
                    'runs': home_runs,
                })

        # Batting rows
        for side, batters in [('away', game.get('visitor_batting', [])),
                              ('home', game.get('home_batting', []))]:
            for b in batters:
                batting_rows.append({
                    'game_id': gid,
                    'team_side': side,
                    'lineup_slot': b.get('lineup_slot', ''),
                    'player_name_jp': b.get('player_name_jp', ''),
                    'position': b.get('position', ''),
                    'is_starter': b.get('is_starter', ''),
                    'sub_type': b.get('sub_type', ''),
                    'ab': b.get('ab', ''),
                    'h': b.get('h', ''),
                    'r': b.get('r', ''),
                    'k': b.get('k', ''),
                    'bb': b.get('bb', ''),
                    'sb': b.get('sb', ''),
                    'e': b.get('e', ''),
                    'hr': b.get('hr', ''),
                })

        # Pitching rows
        for side, pitchers in [('away', game.get('visitor_pitching', [])),
                               ('home', game.get('home_pitching', []))]:
            for p in pitchers:
                pitching_rows.append({
                    'game_id': gid,
                    'team_side': side,
                    'pitch_order': p.get('pitch_order', ''),
                    'player_name_jp': p.get('player_name_jp', ''),
                    'decision': p.get('decision', ''),
                    'ip': p.get('ip', ''),
                    'bf': p.get('bf', ''),
                    'h': p.get('h', ''),
                    'k': p.get('k', ''),
                    'bb': p.get('bb', ''),
                    'er': p.get('er', ''),
                    'e': p.get('e', ''),
                })

        # Home run rows
        for hr in game.get('home_runs', []):
            if 'batter' in hr:
                # Determine team side from team_raw
                # Convert Japanese team name to romanized version for matching
                team_raw_jp = hr.get('team_raw', '')
                team_raw_roman = TEAM_JP_MAP.get(team_raw_jp, team_raw_jp).upper()

                # Get linescore team names (e.g., "RAKUTEN-5" -> "RAKUTEN")
                away_raw = re.sub(r'-\d+$', '', ls.get('away_team', '')).upper()
                home_raw = re.sub(r'-\d+$', '', ls.get('home_team', '')).upper()

                if team_raw_roman and away_raw and team_raw_roman in away_raw:
                    team_side = 'away'
                elif team_raw_roman and home_raw and team_raw_roman in home_raw:
                    team_side = 'home'
                else:
                    team_side = ''

                hr_rows.append({
                    'game_id': gid,
                    'team_side': team_side,
                    'team_raw': team_raw_jp,
                    'batter_name_jp': hr.get('batter', ''),
                    'pitcher_name_jp': hr.get('pitcher', ''),
                    'season_hr_number': hr.get('season_hr_number', ''),
                })

        # Extra-base hit rows
        for xb in game.get('extra_base_hits', []):
            side = 'away' if xb.get('team_side', '').lower() == 'visitor' else 'home'
            # Parse player name to extract count (e.g., "山崎2" -> 2 doubles by Yamazaki)
            parsed = parse_xbh_player(xb.get('player_name_jp', ''))
            player_name = parsed['player_name']
            count = parsed['count']

            # Create one row per hit (so "山崎2" becomes 2 separate rows)
            for _ in range(count):
                xbh_rows.append({
                    'game_id': gid,
                    'team_side': side,
                    'hit_type': xb.get('hit_type', ''),
                    'player_name_jp': player_name,
                })

    # Write CSVs
    def write_csv(path: Path, rows: List[Dict], fieldnames: List[str]):
        with path.open('w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for row in rows:
                writer.writerow({k: row.get(k, '') for k in fieldnames})

    write_csv(output_dir / 'games.csv', games_rows, [
        'game_id', 'year', 'month', 'day', 'date', 'game_number', 'stadium', 'attendance',
        'away_team', 'home_team', 'away_team_raw', 'home_team_raw',
        'away_runs', 'home_runs', 'away_hits', 'home_hits', 'away_errors', 'home_errors',
        'innings_played'
    ])

    write_csv(output_dir / 'linescore.csv', linescore_rows, [
        'game_id', 'team_side', 'inning', 'runs'
    ])

    write_csv(output_dir / 'batting.csv', batting_rows, [
        'game_id', 'team_side', 'lineup_slot', 'player_name_jp', 'position',
        'is_starter', 'sub_type', 'ab', 'h', 'r', 'k', 'bb', 'sb', 'e', 'hr'
    ])

    write_csv(output_dir / 'pitching.csv', pitching_rows, [
        'game_id', 'team_side', 'pitch_order', 'player_name_jp', 'decision',
        'ip', 'bf', 'h', 'k', 'bb', 'er', 'e'
    ])

    write_csv(output_dir / 'home_runs.csv', hr_rows, [
        'game_id', 'team_side', 'team_raw', 'batter_name_jp', 'pitcher_name_jp', 'season_hr_number'
    ])

    write_csv(output_dir / 'extra_base_hits.csv', xbh_rows, [
        'game_id', 'team_side', 'hit_type', 'player_name_jp'
    ])

    print(f"Wrote {len(games_rows)} games to {output_dir}")
    print(f"  - linescore.csv: {len(linescore_rows)} rows")
    print(f"  - batting.csv: {len(batting_rows)} rows")
    print(f"  - pitching.csv: {len(pitching_rows)} rows")
    print(f"  - home_runs.csv: {len(hr_rows)} rows")
    print(f"  - extra_base_hits.csv: {len(xbh_rows)} rows")


def main():
    parser = argparse.ArgumentParser(description='Convert NPB boxscores to CSV')
    parser.add_argument('--input-dir', required=True, help='Directory containing XLSX boxscores')
    parser.add_argument('--output-dir', required=True, help='Directory for CSV output')
    parser.add_argument('--glob', default='*.xlsx', help='File pattern (default: *.xlsx)')
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)

    xlsx_files = sorted(input_dir.glob(args.glob))

    if not xlsx_files:
        print(f"No files matching '{args.glob}' found in {input_dir}")
        return

    print(f"Processing {len(xlsx_files)} boxscore files...")

    games_data = []
    for xlsx_path in xlsx_files:
        try:
            game = convert_boxscore(xlsx_path)
            games_data.append(game)
            print(f"  Parsed: {xlsx_path.name}")
        except Exception as e:
            print(f"  ERROR parsing {xlsx_path.name}: {e}")

    if games_data:
        write_csvs(games_data, output_dir)


if __name__ == '__main__':
    main()
